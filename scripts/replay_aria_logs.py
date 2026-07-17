"""Replay ARIA conversation log prompts against our API and compare token usage."""

from __future__ import annotations

import json
import os
import uuid
from pathlib import Path

import httpx
from openpyxl import load_workbook

LOGS_DIR = Path(__file__).resolve().parent.parent / "AIConversationLogs"
OUTPUT_FILE = Path(__file__).resolve().parent.parent / "logs_replay_results.json"
BASE_URL = os.environ.get("SMOKE_BASE_URL", "http://127.0.0.1:9200")


def extract_prompts() -> list[dict]:
  prompts: list[dict] = []
  for f in sorted(LOGS_DIR.glob("AI_Conversation_Log_Detail_*.xlsx")):
    wb = load_workbook(f, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    prompt = None
    total_tokens = None
    for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
      if row and row[0]:
        key = str(row[0]).strip().lower()
        if key == "prompt":
          prompt = str(row[1]) if row[1] else ""
        elif key == "total tokens":
          total_tokens = row[1]
    wb.close()
    if prompt:
      prompts.append({
        "file": f.name,
        "prompt": prompt,
        "aria_total_tokens": int(str(total_tokens).replace(",", "")) if total_tokens else None,
      })
  return prompts


def replay_mode(client: httpx.Client, mode: str, prompts: list[dict]) -> list[dict]:
  results: list[dict] = []
  user = client.post("/users", json={"email": f"replay+{mode}+{uuid.uuid4().hex[:8]}@uni.edu"}).json()

  if mode == "same_session":
    session = client.post("/sessions", json={"user_id": user["id"], "title": f"ARIA Replay ({mode})"}).json()
    sessions = [session["id"]] * len(prompts)
  else:
    sessions = []
    for _ in prompts:
      s = client.post("/sessions", json={"user_id": user["id"], "title": "ARIA Fresh Session"}).json()
      sessions.append(s["id"])

  for i, item in enumerate(prompts):
    resp = client.post("/chat", json={
      "user_id": user["id"],
      "session_id": sessions[i],
      "message": item["prompt"],
    }, timeout=300.0).json()

    results.append({
      "file": item["file"],
      "mode": mode,
      "turn": i + 1,
      "prompt_preview": item["prompt"][:100],
      "aria_total_tokens": item["aria_total_tokens"],
      "our_context_tokens": resp["context_tokens_used"],
      "our_naive_baseline": resp["naive_baseline_tokens"],
      "savings_percent": resp["savings_percent"],
      "retrieved_count": len(resp.get("retrieved_memories", [])),
      "latency_ms": resp.get("latency_ms"),
      "has_session_summary": resp.get("has_session_summary"),
    })

  return results


def main() -> None:
  prompts = extract_prompts()
  if not prompts:
    raise SystemExit("No prompts found in AIConversationLogs")

  with httpx.Client(base_url=BASE_URL, timeout=300.0) as client:
    health = client.get("/health")
    health.raise_for_status()

    same_session = replay_mode(client, "same_session", prompts)
    fresh_session = replay_mode(client, "fresh_session", prompts)

  all_results = same_session + fresh_session

  aria_total = sum(p["aria_total_tokens"] or 0 for p in prompts)
  our_total = sum(r["our_context_tokens"] for r in same_session)
  savings = round((1 - our_total / aria_total) * 100, 2) if aria_total else 0

  output = {
    "summary": {
      "prompts_replayed": len(prompts),
      "aria_total_tokens": aria_total,
      "our_total_context_tokens_same_session": our_total,
      "overall_savings_vs_aria_percent": savings,
    },
    "results": all_results,
  }

  OUTPUT_FILE.write_text(json.dumps(output, indent=2), encoding="utf-8")
  print(json.dumps(output["summary"], indent=2))
  print(f"Wrote {OUTPUT_FILE}")


if __name__ == "__main__":
  main()
