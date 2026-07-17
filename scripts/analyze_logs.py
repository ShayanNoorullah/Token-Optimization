"""Parse AIConversationLogs xlsx files and generate logs_issue.txt."""

from __future__ import annotations

import os
from pathlib import Path

from openpyxl import load_workbook

LOGS_DIR = Path(__file__).resolve().parent.parent / "AIConversationLogs"
OUTPUT_FILE = Path(__file__).resolve().parent.parent / "logs_issue.txt"


def _read_field_value(ws, field_name: str) -> str | None:
    for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
        if row and row[0] and str(row[0]).strip().lower() == field_name.lower():
            return str(row[1]) if len(row) > 1 and row[1] is not None else None
    return None


def parse_detail_file(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    summary_ws = wb[wb.sheetnames[0]]

    info = {
        "file": path.name,
        "prompt": _read_field_value(summary_ws, "Prompt") or "",
        "response": (_read_field_value(summary_ws, "Response") or "")[:300],
        "total_tokens": _read_field_value(summary_ws, "Total Tokens"),
        "response_time_ms": _read_field_value(summary_ws, "Response Time (ms)"),
        "sql_queries": _read_field_value(summary_ws, "SQL Queries"),
        "status": _read_field_value(summary_ws, "Status"),
        "llm_calls": [],
        "sql_tools": [],
    }

    if len(wb.sheetnames) > 1:
        token_ws = wb[wb.sheetnames[1]]
        for row in token_ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None:
                info["llm_calls"].append({
                    "type": row[1],
                    "prompt_tokens": row[4],
                    "completion_tokens": row[5],
                    "total": row[6],
                    "time_ms": row[7],
                })

    if len(wb.sheetnames) > 2:
        sql_ws = wb[wb.sheetnames[2]]
        for row in sql_ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                info["sql_tools"].append({
                    "tool": row[0],
                    "status": row[1],
                    "rows": row[3],
                })

    wb.close()
    return info


def parse_summary_file(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    conversations = []
    for row in rows:
        if not row or not row[4]:
            continue
        conversations.append({
            "datetime": row[0],
            "user": row[1],
            "prompt": str(row[4])[:200],
            "prompt_tokens": row[6],
            "completion_tokens": row[7],
            "total_tokens": row[8],
            "response_time_ms": row[9],
            "status": row[10],
        })
    return conversations


def analyze_all() -> tuple[list[dict], list[dict]]:
    details = []
    for f in sorted(LOGS_DIR.glob("AI_Conversation_Log_Detail_*.xlsx")):
        details.append(parse_detail_file(f))

    summaries = []
    for f in sorted(LOGS_DIR.glob("AI_Conversation_Logs_*.xlsx")):
        summaries = parse_summary_file(f)

    return details, summaries


def build_issues_report(details: list[dict], summaries: list[dict]) -> str:
    lines = [
        "ARIA Conversation Log Issues",
        "=" * 50,
        "",
        "This file lists problems found in the AIConversationLogs folder.",
        "These logs come from a similar AI chat system (ARIA / BullseyeEZ).",
        "Our token-efficient system is designed to fix these problems.",
        "",
    ]

    token_counts = []
    for d in details:
        try:
            token_counts.append(int(str(d.get("total_tokens", "0")).replace(",", "")))
        except ValueError:
            pass

    briefing_count = sum(
        1 for d in details
        if "briefing" in (d.get("prompt") or "").lower()
        or "dashboard" in (d.get("prompt") or "").lower()
    )

    multi_call_count = sum(1 for d in details if len(d.get("llm_calls", [])) > 1)
    sql_heavy = [d for d in details if d.get("sql_tools")]

    large_files = [
        d["file"] for d in details
        if (LOGS_DIR / d["file"]).stat().st_size > 50000
    ]

    issues = [
        (
            "1. Too many tokens per message",
            f"Most chats use {min(token_counts) if token_counts else '?'} to "
            f"{max(token_counts) if token_counts else '?'} tokens per request. "
            f"SQL questions spike to around 60,000 tokens. "
            f"Our system targets 100-1,500 context tokens instead.",
        ),
        (
            "2. One user question triggers multiple AI calls",
            f"{multi_call_count} of {len(details)} conversations used more than one "
            f"LLM call (InitialPrompt + SqlGeneration + FinalResponse). "
            f"This triples cost and latency for a single user message.",
        ),
        (
            "3. Same briefing asked repeatedly with no memory reuse",
            f"{briefing_count} conversations are dashboard briefing requests. "
            f"Each one resends the full context instead of reusing a stored summary.",
        ),
        (
            "4. Huge log files from storing full API payloads",
            f"Files over 50KB: {', '.join(large_files) or 'none'}. "
            f"Full Provider Request/Response JSON is stored, making logs hard to read "
            f"and wasting storage.",
        ),
        (
            "5. SQL query results dumped raw into AI context",
            f"{len(sql_heavy)} conversations ran SQL tools returning 16-60 rows of data "
            f"directly into the prompt. Large tabular data inflates tokens fast.",
        ),
        (
            "6. AI answered before data was ready",
            "In Detail 7, the dashboard was still showing 'Processing...' but the AI "
            "still generated a full briefing. Responses can be based on empty or stale data.",
        ),
        (
            "7. Slow responses on heavy queries",
            "SQL-heavy conversations (Detail 4: ~59K tokens, Detail 6: ~62K tokens) "
            "take much longer than text-only briefings (~6K tokens).",
        ),
        (
            "8. No selective memory retrieval",
            "Every call appears to resend the entire conversation context. "
            f"There is no evidence of semantic retrieval or summarization to cut token use.",
        ),
        (
            "9. Questionable or confusing metrics in responses",
            "Some responses report metrics over 100% (e.g. lock rate 124.19%). "
            "This suggests context confusion, bad data, or hallucination.",
        ),
    ]

    for title, description in issues:
        lines.extend([title, "-" * len(title), description, ""])

    lines.extend([
        "Evidence Summary",
        "=" * 50,
        "",
    ])

    for d in details:
        lines.append(f"File: {d['file']}")
        lines.append(f"  Prompt: {d['prompt'][:120]}...")
        lines.append(f"  Total tokens: {d.get('total_tokens', 'N/A')}")
        lines.append(f"  LLM calls: {len(d.get('llm_calls', []))}")
        lines.append(f"  SQL tools: {len(d.get('sql_tools', []))}")
        lines.append(f"  Response time: {d.get('response_time_ms', 'N/A')} ms")
        lines.append("")

    if summaries:
        lines.extend(["Summary Index (all conversations)", "-" * 40, ""])
        for i, s in enumerate(summaries, 1):
            lines.append(
                f"  {i}. {s.get('total_tokens', '?')} tokens | "
                f"{s.get('response_time_ms', '?')} ms | "
                f"{s.get('prompt', '')[:80]}..."
            )

    return "\n".join(lines)


def main() -> None:
    if not LOGS_DIR.exists():
        raise SystemExit(f"Logs directory not found: {LOGS_DIR}")

    details, summaries = analyze_all()
    report = build_issues_report(details, summaries)
    OUTPUT_FILE.write_text(report, encoding="utf-8")
    print(f"Wrote {OUTPUT_FILE} ({len(details)} detail files, {len(summaries)} summary rows)")


if __name__ == "__main__":
    main()
